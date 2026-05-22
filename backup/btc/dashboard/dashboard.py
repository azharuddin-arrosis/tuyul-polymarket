import os
import sys
import time
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict
from math import floor
import threading

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


@dataclass
class TradeRecord:
    id: int
    timestamp: str
    window_ts: int
    direction: str
    confidence: float
    bet_size: float
    token_price: float
    shares: int
    actual_direction: str
    win: bool
    pnl: float
    bankroll: float
    score: float


@dataclass
class DashboardState:
    start_time: str
    start_bankroll: float
    current_bankroll: float
    mode: str
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    win_rate: float = 0.0
    total_pnl: float = 0.0
    roi_pct: float = 0.0
    current_price: float = 0.0
    window_open_price: float = 0.0
    delta_pct: float = 0.0
    signal_direction: str = "NEUTRAL"
    signal_confidence: float = 0.0
    next_window: int = 0
    status: str = "IDLE"


class ExcelDashboard:
    def __init__(self, output_path: str = "dashboard.xlsx"):
        self.output_path = output_path
        self.wb = Workbook()
        self.state = DashboardState(
            start_time=datetime.now(timezone.utc).isoformat(),
            start_bankroll=100.0,
            current_bankroll=100.0,
            mode="safe"
        )
        self.trades: List[TradeRecord] = []
        self.lock = threading.Lock()
        self._setup_workbook()
    
    def _setup_workbook(self):
        self.wb.remove(self.wb.active)
        
        self.ws_summary = self.wb.create_sheet("Summary", 0)
        self.ws_live = self.wb.create_sheet("Live Signal", 1)
        self.ws_trades = self.wb.create_sheet("Trade History", 2)
        self.ws_curve = self.wb.create_sheet("Bankroll Curve", 3)
        
        self._setup_summary_sheet()
        self._setup_live_sheet()
        self._setup_trades_sheet()
        self._setup_curve_sheet()
    
    def _setup_summary_sheet(self):
        ws = self.ws_summary
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells("A1:D1")
        ws["A1"] = "POLYMARKET BTC 5-MINUTE BOT"
        ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws["A1"].alignment = center
        
        ws["A3"] = "SESSION"
        ws["B3"] = self.state.start_time
        ws["A4"] = "MODE"
        ws["B4"] = self.state.mode
        ws["A5"] = "STATUS"
        ws["B5"] = self.state.status
        
        labels = [
            ("A7", "METRICS"),
            ("A8", "Start Bankroll"),
            ("A9", "Current Bankroll"),
            ("A10", "Total PnL"),
            ("A11", "ROI"),
            ("A12", "Win Rate"),
            ("A14", "TRADES"),
            ("A15", "Total Trades"),
            ("A16", "Wins"),
            ("A17", "Losses"),
        ]
        
        for cell, val in labels:
            ws[cell] = val
            ws[cell].font = Font(bold=True, size=10)
        
        for cell in ["B8", "B9", "B10", "B11", "B12", "B15", "B16", "B17"]:
            ws[cell].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _setup_live_sheet(self):
        ws = self.ws_live
        
        ws.merge_cells("A1:E1")
        ws["A1"] = "LIVE TRADING SIGNAL"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        
        header_fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        headers = ["Metric", "Value"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col).value = h
            ws.cell(row=2, column=col).fill = header_fill
            ws.cell(row=2, column=col).font = header_font
        
        metrics = [
            ("BTC Price", "A3"),
            ("Window Open", "A4"),
            ("Delta %", "A5"),
            ("Signal", "A7"),
            ("Confidence", "A8"),
            ("Next Window", "A10"),
            ("Seconds Left", "A11"),
        ]
        
        for _, cell in metrics:
            ws[cell].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 14
    
    def _setup_trades_sheet(self):
        ws = self.ws_trades
        
        ws.merge_cells("A1:M1")
        ws["A1"] = "TRADE HISTORY"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=9)
        
        cols = ["#", "Time", "Window", "Direction", "Conf", "Bet", "Price", "Shares", "Actual", "Win", "PnL", "Bankroll", "Score"]
        
        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for col in range(1, 14):
            ws.column_dimensions[get_column_letter(col)].width = 10
        
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
    
    def _setup_curve_sheet(self):
        ws = self.ws_curve
        
        ws["A1"] = "BANKROLL CURVE"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for col, h in enumerate(["Trade #", "Bankroll"], 1):
            ws.cell(row=2, column=col).value = h
            ws.cell(row=2, column=col).fill = header_fill
            ws.cell(row=2, column=col).font = header_font
        
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 14
    
    def update_state(self, state: DashboardState):
        with self.lock:
            self.state = state
            self._refresh_summary()
            self._refresh_live()
    
    def add_trade(self, trade: TradeRecord):
        with self.lock:
            self.trades.append(trade)
            self._add_trade_row(trade)
            self._add_curve_point(trade)
            self._refresh_summary()
    
    def _refresh_summary(self):
        ws = self.ws_summary
        
        ws["B3"] = self.state.start_time
        ws["B4"] = self.state.mode
        ws["B5"] = self.state.status
        
        ws["B8"] = f"${self.state.start_bankroll:.2f}"
        ws["B9"] = f"${self.state.current_bankroll:.2f}"
        ws["B10"] = f"${self.state.total_pnl:.2f}"
        ws["B11"] = f"{self.state.roi_pct:.1f}%"
        ws["B12"] = f"{self.state.win_rate:.1%}"
        
        ws["B15"] = self.state.total_trades
        ws["B16"] = self.state.wins
        ws["B17"] = self.state.losses
        
        if self.state.total_pnl > 0:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws["B10"].fill = fill
        
        if self.state.roi_pct > 0:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws["B11"].fill = fill
    
    def _refresh_live(self):
        ws = self.ws_live
        
        ws["B3"] = f"${self.state.current_price:.2f}"
        ws["B4"] = f"${self.state.window_open_price:.2f}"
        
        delta_cell = ws["B5"]
        delta_cell.value = f"{self.state.delta_pct:.3f}%"
        
        if self.state.delta_pct > 0:
            delta_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif self.state.delta_pct < 0:
            delta_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        signal_cell = ws["B7"]
        signal_cell.value = self.state.signal_direction.upper()
        
        if self.state.signal_direction.upper() == "UP":
            signal_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            signal_cell.font = Font(bold=True, size=12, color="006100")
        elif self.state.signal_direction.upper() == "DOWN":
            signal_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            signal_cell.font = Font(bold=True, size=12, color="9C0006")
        else:
            signal_cell.font = Font(bold=True, size=12)
        
        ws["B8"] = f"{self.state.signal_confidence:.0%}"
        ws["B10"] = datetime.fromtimestamp(self.state.next_window).strftime("%H:%M:%S")
        
        seconds_left = max(0, self.state.next_window - int(time.time()))
        ws["B11"] = f"{seconds_left}s"
        
        self.save()
    
    def _add_trade_row(self, trade: TradeRecord):
        ws = self.ws_trades
        
        row = len(self.trades) + 3
        
        ws.cell(row=row, column=1).value = trade.id
        ws.cell(row=row, column=2).value = trade.timestamp
        ws.cell(row=row, column=3).value = datetime.fromtimestamp(trade.window_ts).strftime("%m/%d %H:%M")
        ws.cell(row=row, column=4).value = trade.direction.upper()
        ws.cell(row=row, column=5).value = f"{trade.confidence:.0%}"
        ws.cell(row=row, column=6).value = f"${trade.bet_size:.2f}"
        ws.cell(row=row, column=7).value = f"${trade.token_price:.2f}"
        ws.cell(row=row, column=8).value = trade.shares
        ws.cell(row=row, column=9).value = trade.actual_direction.upper()
        ws.cell(row=row, column=10).value = "YES" if trade.win else "NO"
        ws.cell(row=row, column=11).value = f"${trade.pnl:.2f}"
        ws.cell(row=row, column=12).value = f"${trade.bankroll:.2f}"
        ws.cell(row=row, column=13).value = f"{trade.score:.2f}"
        
        if trade.win:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col in range(1, 14):
            ws.cell(row=row, column=col).fill = fill
    
    def _add_curve_point(self, trade: TradeRecord):
        ws = self.ws_curve
        
        row = trade.id + 2
        
        ws.cell(row=row, column=1).value = trade.id
        ws.cell(row=row, column=2).value = trade.bankroll
    
    def save(self):
        self.wb.save(self.output_path)
    
    def close(self):
        self.save()
        print(f"Dashboard saved to {self.output_path}")


def create_dashboard(output_path: str = "polymarket_dashboard.xlsx") -> ExcelDashboard:
    return ExcelDashboard(output_path)


def main():
    dashboard = create_dashboard()
    print(f"Created dashboard: {dashboard.output_path}")
    
    dashboard.state.current_price = 45000.00
    dashboard.state.window_open_price = 44980.00
    dashboard.state.delta_pct = 0.044
    dashboard.state.signal_direction = "UP"
    dashboard.state.signal_confidence = 0.75
    dashboard.state.current_bankroll = 150.00
    dashboard.state.total_pnl = 50.00
    dashboard.state.roi_pct = 50.0
    dashboard.state.win_rate = 0.65
    dashboard.state.total_trades = 20
    dashboard.state.wins = 13
    dashboard.state.losses = 7
    
    dashboard.update_state(dashboard.state)
    
    for i in range(1, 6):
        trade = TradeRecord(
            id=i,
            timestamp=datetime.now(timezone.utc).isoformat(),
            window_ts=int(time.time()) - (i * 300),
            direction="up" if i % 2 == 1 else "down",
            confidence=0.7,
            bet_size=25.0,
            token_price=0.50,
            shares=50,
            actual_direction="up" if i % 2 == 1 else "down",
            win=True,
            pnl=25.0,
            bankroll=100.0 + (i * 25.0),
            score=3.5
        )
        dashboard.add_trade(trade)
    
    dashboard.close()


if __name__ == "__main__":
    main()