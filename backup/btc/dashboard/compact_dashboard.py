import os
import sys
import time
import json
import asyncio
import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Optional, Callable
from dataclasses import dataclass, field, asdict
from math import floor
from threading import Thread
import logging
import traceback

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.formatting.rule import DataBarRule
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/dashboard.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class TradeRecord:
    id: int
    timestamp: str
    window_ts: int
    direction: str
    confidence: float
    bet_size: float
    token_price: float
    shares: int
    actual_direction: str
    win: bool
    pnl: float
    bankroll: float
    score: float
    strategy_breakdown: Dict = field(default_factory=dict)


@dataclass
class LiveMetrics:
    current_price: float = 0.0
    window_open_price: float = 0.0
    delta_pct: float = 0.0
    signal_direction: str = "NEUTRAL"
    signal_confidence: float = 0.0
    signal_score: float = 0.0
    next_window_ts: int = 0
    seconds_until_close: int = 0
    last_update: str = ""
    rsi_14: float = 50.0
    ema_9: float = 0.0
    ema_21: float = 0.0
    volume_ratio: float = 1.0


@dataclass
class DashboardState:
    mode: str = "safe"
    start_time: str = ""
    start_bankroll: float = 0.0
    current_bankroll: float = 0.0
    total_pnl: float = 0.0
    roi_pct: float = 0.0
    win_rate: float = 0.0
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    max_drawdown: float = 0.0
    streak_current: int = 0
    streak_best: int = 0
    status: str = "IDLE"
    last_error: str = ""
    live: LiveMetrics = field(default_factory=LiveMetrics)


class CompactExcelDashboard:
    COMPACT_HEADER = NamedStyle(name="compact_header")
    COMPACT_HEADER.font = Font(bold=True, size=8, color="FFFFFF")
    COMPACT_HEADER.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
    COMPACT_HEADER.alignment = Alignment(horizontal="center", vertical="center")
    
    CELL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    CELL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    CELL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    CELL_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    def __init__(self, output_path: str = "polymarket_dashboard.xlsx"):
        self.output_path = output_path
        self.wb = Workbook(iso_datesheet=True)
        self.state = DashboardState(
            start_time=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            start_bankroll=100.0,
            current_bankroll=100.0,
            mode="safe"
        )
        self.trades: List[TradeRecord] = []
        self.lock = threading.RLock()
        self.running = False
        self._setup_styles()
        self._setup_workbook()
        self._save_thread: Optional[Thread] = None
        self._save_interval = 5
    
    def _setup_styles(self):
        try:
            self.wb.add_named_style(self.COMPACT_HEADER)
        except:
            pass
    
    def _setup_workbook(self):
        self.wb.remove(self.wb.active)
        
        self.ws_dashboard = self.wb.create_sheet("DASHBOARD", 0)
        self.ws_trades = self.wb.create_sheet("TRADES", 1)
        self.ws_curve = self.wb.create_sheet("CURVE", 2)
        self.ws_live = self.wb.create_sheet("LIVE", 3)
        self.ws_log = self.wb.create_sheet("LOG", 4)
        
        self._build_dashboard_sheet()
        self._build_trades_sheet()
        self._build_curve_sheet()
        self._build_live_sheet()
        self._build_log_sheet()
    
    def _build_dashboard_sheet(self):
        ws = self.ws_dashboard
        ws.freeze_panes = "A3"
        
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12
        
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "POLYMARKET BTC 5-MINUTE TRADING BOT"
        cell.font = Font(bold=True, size=14, color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        
        ws["A2"] = "STATUS"
        ws["B2"] = self.state.status
        ws["D2"] = self.state.start_time
        
        for cell in ["A2", "B2", "D2"]:
            ws[cell].font = Font(bold=True, size=9)
        
        meta_pairs = [
            ("A4", "MODE", "B4", "safe"),
            ("A5", "START", "B5", "$100.00"),
            ("A6", "CURRENT", "B6", "$100.00"),
            ("A7", "P&L", "B7", "$0.00"),
            ("A8", "ROI", "B8", "0.0%"),
        ]
        
        stat_pairs = [
            ("C4", "TRADES", "D4", "0"),
            ("C5", "WINS", "D5", "0"),
            ("C6", "LOSSES", "D6", "0"),
            ("C7", "WIN RATE", "D7", "0.0%"),
            ("C8", "MAX DD", "D8", "0.0%"),
        ]
        
        for c1, v1, c2, v2 in meta_pairs + stat_pairs:
            ws[c1] = v1
            ws[c2] = v2
            
        for range_prefix in ["A", "C"]:
            for row in range(4, 9):
                ws[f"{range_prefix}{row}"].font = Font(size=9, bold=True)
                ws[f"{range_prefix}{row}"].fill = self.CELL_GRAY
                ws[get_column_letter(ord(range_prefix) - ord("A") + 2) + str(row)].font = Font(size=9)
        
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for row in range(3, 9):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                if not cell.border:
                    cell.border = border
    
    def _build_trades_sheet(self):
        ws = self.ws_trades
        ws.freeze_panes = "A3"
        
        headers = ["#", "TIME", "WINDOW", "DIR", "CONF", "BET", "TOKEN", "SH", "ACTUAL", "WIN", "PNL", "BANKROLL", "SCORE"]
        widths = [5, 12, 10, 6, 6, 8, 7, 5, 7, 5, 8, 10, 7]
        
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = w
        
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).font = Font(size=7)
    
    def _build_curve_sheet(self):
        ws = self.ws_curve
        
        for col, h in enumerate(["#", "BANKROLL", "PNL"], 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
    
    def _build_live_sheet(self):
        ws = self.ws_live
        
        ws["A1"] = "LIVE METRICS"
        ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
        
        metrics = [
            ("BTC Price", "A3"),
            ("Window Open", "A4"),
            ("Delta %", "A5"),
            ("RSI(14)", "A6"),
            ("EMA9", "A7"),
            ("EMA21", "A8"),
            ("Vol Ratio", "A9"),
            ("", "A10"),
            ("SIGNAL", "A11"),
            ("Direction", "A12"),
            ("Confidence", "A13"),
            ("Score", "A14"),
            ("", "A15"),
            ("NEXT", "A16"),
            ("Window", "A17"),
            ("Countdown", "A18"),
        ]
        
        for _, cell_ref in metrics:
            ws[cell_ref].font = Font(size=9, bold=True)
            ws[cell_ref].fill = self.CELL_GRAY
        
        ws.column_dimensions["A"].width = 14
        for col in ["B"]:
            ws.column_dimensions[col].width = 14
    
    def _build_log_sheet(self):
        ws = self.ws_log
        
        ws["A1"] = "ACTIVITY LOG"
        ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
        
        for col, h in enumerate(["TIME", "LEVEL", "MESSAGE"], 1):
            cell = ws.cell(row=2, column=col)
            cell.value = h
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
        
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 60
    
    def update_state(self, mode: str = None, status: str = None, start_bankroll: float = None,
                   current_bankroll: float = None):
        with self.lock:
            if mode:
                self.state.mode = mode
            if status:
                self.state.status = status
            if start_bankroll is not None:
                self.state.start_bankroll = start_bankroll
            if current_bankroll is not None:
                self.state.current_bankroll = current_bankroll
            
            self.state.total_pnl = self.state.current_bankroll - self.state.start_bankroll
            self.state.roi_pct = (self.state.total_pnl / self.state.start_bankroll * 100) if self.state.start_bankroll > 0 else 0
            
            if self.state.total_trades > 0:
                self.state.win_rate = self.state.wins / self.state.total_trades
            
            self._render_dashboard()
    
    def update_live_metrics(self, metrics: LiveMetrics):
        with self.lock:
            self.state.live = metrics
            self._render_live()
    
    def add_trade(self, trade: TradeRecord):
        with self.lock:
            self.trades.append(trade)
            
            self.state.total_trades += 1
            if trade.win:
                self.state.wins += 1
                self.state.streak_current += 1
            else:
                self.state.losses += 1
                self.state.streak_current = 0
            
            self.state.streak_best = max(self.state.streak_best, self.state.streak_current)
            
            peak = max(t.bankroll for t in self.trades)
            current = self.state.current_bankroll
            drawdown = ((peak - current) / peak * 100) if peak > 0 else 0
            self.state.max_drawdown = max(self.state.max_drawdown, drawdown)
            
            self._render_trade_row(trade)
            self._render_curve_row(trade)
            self._render_dashboard()
            
            self._schedule_save()
    
    def log_event(self, message: str, level: str = "INFO"):
        with self.lock:
            ws = self.ws_log
            row = 3
            
            while ws.cell(row=row, column=1).value:
                row += 1
            
            ws.cell(row=row, column=1).value = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row, column=2).value = level
            ws.cell(row=row, column=3).value = message
            
            if level == "ERROR":
                ws.cell(row=row, column=2).fill = self.CELL_RED
            elif level == "WARNING":
                ws.cell(row=row, column=2).fill = self.CELL_YELLOW
            
            if row > 1000:
                for r in range(3, row - 100):
                    ws.delete_rows(3)
    
    def _render_dashboard(self):
        ws = self.ws_dashboard
        
        ws["B2"] = self.state.status
        
        ws["B4"] = self.state.mode
        ws["B5"] = f"${self.state.start_bankroll:.2f}"
        ws["B6"] = f"${self.state.current_bankroll:.2f}"
        
        pnl_cell = ws["B7"]
        pnl_cell.value = f"${self.state.total_pnl:+.2f}"
        if self.state.total_pnl >= 0:
            pnl_cell.fill = self.CELL_GREEN
        else:
            pnl_cell.fill = self.CELL_RED
        
        roi_cell = ws["B8"]
        roi_cell.value = f"{self.state.roi_pct:+.1f}%"
        if self.state.roi_pct >= 0:
            roi_cell.fill = self.CELL_GREEN
        else:
            roi_cell.fill = self.CELL_RED
        
        ws["D4"] = str(self.state.total_trades)
        ws["D5"] = str(self.state.wins)
        ws["D6"] = str(self.state.losses)
        
        wr_cell = ws["D7"]
        wr_cell.value = f"{self.state.win_rate:.1%}"
        if self.state.win_rate >= 0.55:
            wr_cell.fill = self.CELL_GREEN
        elif self.state.win_rate >= 0.45:
            wr_cell.fill = self.CELL_YELLOW
        else:
            wr_cell.fill = self.CELL_RED
        
        ws["D8"] = f"-{self.state.max_drawdown:.1f}%"
        
        self.save_quiet()
    
    def _render_live(self):
        ws = self.ws_live
        
        m = self.state.live
        
        ws["B3"] = f"${m.current_price:,.2f}"
        ws["B4"] = f"${m.window_open_price:,.2f}"
        
        delta_cell = ws["B5"]
        delta_cell.value = f"{m.delta_pct:+.3f}%"
        if m.delta_pct > 0:
            delta_cell.fill = self.CELL_GREEN
        elif m.delta_pct < 0:
            delta_cell.fill = self.CELL_RED
        
        ws["B6"] = f"{m.rsi_14:.1f}"
        ws["B7"] = f"${m.ema_9:,.2f}"
        ws["B8"] = f"${m.ema_21:,.2f}"
        ws["B9"] = f"{m.volume_ratio:.2f}x"
        
        dir_cell = ws["B12"]
        dir_cell.value = m.signal_direction.upper()
        if m.signal_direction.upper() == "UP":
            dir_cell.fill = self.CELL_GREEN
            dir_cell.font = Font(bold=True, size=11)
        elif m.signal_direction.upper() == "DOWN":
            dir_cell.fill = self.CELL_RED
            dir_cell.font = Font(bold=True, size=11)
        
        ws["B13"] = f"{m.signal_confidence:.0%}"
        ws["B14"] = f"{m.signal_score:.2f}"
        
        if m.next_window_ts > 0:
            ws["B17"] = datetime.fromtimestamp(m.next_window_ts).strftime("%H:%M:%S")
            ws["B18"] = f"{m.seconds_until_close}s"
        
        self.save_quiet()
    
    def _render_trade_row(self, trade: TradeRecord):
        ws = self.ws_trades
        row = trade.id + 2
        
        ws.cell(row=row, column=1).value = trade.id
        ws.cell(row=row, column=2).value = trade.timestamp
        ws.cell(row=row, column=3).value = datetime.fromtimestamp(trade.window_ts).strftime("%m/%d %H:%M")
        ws.cell(row=row, column=4).value = trade.direction.upper()
        ws.cell(row=row, column=5).value = f"{trade.confidence:.0%}"
        ws.cell(row=row, column=6).value = f"${trade.bet_size:.2f}"
        ws.cell(row=row, column=7).value = f"${trade.token_price:.2f}"
        ws.cell(row=row, column=8).value = trade.shares
        ws.cell(row=row, column=9).value = trade.actual_direction.upper()
        ws.cell(row=row, column=10).value = "WIN" if trade.win else "LOSS"
        ws.cell(row=row, column=11).value = f"${trade.pnl:+.2f}"
        ws.cell(row=row, column=12).value = f"${trade.bankroll:.2f}"
        ws.cell(row=row, column=13).value = f"{trade.score:.2f}"
        
        fill = self.CELL_GREEN if trade.win else self.CELL_RED
        
        for col in range(1, 14):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).font = Font(size=8)
    
    def _render_curve_row(self, trade: TradeRecord):
        ws = self.ws_curve
        
        row = trade.id + 1
        
        ws.cell(row=row, column=1).value = trade.id
        ws.cell(row=row, column=2).value = trade.bankroll
        ws.cell(row=row, column=3).value = trade.pnl
    
    def _schedule_save(self):
        if self._save_thread and self._save_thread.is_alive():
            return
        
        self._save_thread = Thread(target=self._delayed_save, daemon=True)
        self._save_thread.start()
    
    def _delayed_save(self):
        time.sleep(self._save_interval)
        self.save_quiet()
    
    def save_quiet(self):
        try:
            self.wb.save(self.output_path)
        except Exception as e:
            logger.warning(f"Quiet save failed: {e}")
    
    def save(self):
        self.wb.save(self.output_path)
        logger.info(f"Dashboard saved: {self.output_path}")
    
    def close(self):
        self.running = False
        self.save()
        logger.info("Dashboard closed")


def create_dashboard(output_path: str = "polymarket_dashboard.xlsx") -> CompactExcelDashboard:
    return CompactExcelDashboard(output_path)


def test_dashboard():
    db = create_dashboard("test_dashboard.xlsx")
    
    db.update_state(mode="safe", status="RUNNING", start_bankroll=100.0, current_bankroll=100.0)
    
    db.update_live_metrics(LiveMetrics(
        current_price=45000.00,
        window_open_price=44950.00,
        delta_pct=0.111,
        signal_direction="UP",
        signal_confidence=0.75,
        signal_score=5.2,
        next_window_ts=int(time.time()) + 120,
        seconds_until_close=120,
        rsi_14=62.5,
        ema_9=44980.0,
        ema_21=44920.0,
        volume_ratio=1.3
    ))
    
    for i in range(1, 21):
        trade = TradeRecord(
            id=i,
            timestamp=datetime.now(timezone.utc).isoformat(),
            window_ts=int(time.time()) - (21 - i) * 300,
            direction="up" if i % 3 != 0 else "down",
            confidence=0.6 + (i % 3) * 0.1,
            bet_size=25.0,
            token_price=0.50,
            shares=50,
            actual_direction="up" if i % 3 != 0 else "down",
            win=i % 3 != 0,
            pnl=25.0 if i % 3 != 0 else -25.0,
            bankroll=100.0 + (i * 16),
            score=3.5 + (i % 3)
        )
        db.add_trade(trade)
    
    db.log_event("Test trade executed", "INFO")
    db.close()
    
    print(f"Test dashboard created: test_dashboard.xlsx")


if __name__ == "__main__":
    os.makedirs("logs", exist_ok=True)
    test_dashboard()