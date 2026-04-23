import argparse
import json
from datetime import datetime, timezone
from typing import List, Dict
from dataclasses import dataclass
from math import floor
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backtest import fetch_historical_data
from strategy import analyze


@dataclass
class BacktestResult:
    config: str
    confidence_threshold: float
    mode: str
    total_trades: int
    wins: int
    losses: int
    win_rate: float
    final_bankroll: float
    total_pnl: float
    max_drawdown: float
    trades: List[Dict]


def simulate_token_price(delta_pct: float) -> float:
    abs_delta = abs(delta_pct)
    if abs_delta < 0.005:
        return 0.50
    elif abs_delta < 0.02:
        return 0.55
    elif abs_delta < 0.05:
        return 0.65
    elif abs_delta < 0.10:
        return 0.80
    return min(0.95, 0.92 + (abs_delta - 0.10) * 0.3)


def run_backtest(
    candles: List[Dict],
    confidence_thresholds: List[float],
    modes: List[str],
    starting_bankroll: float = 1.0,
    min_bet: float = 1.0
) -> List[BacktestResult]:
    results = []
    
    window_size = 5 * 60
    bankroll = starting_bankroll
    
    for conf_thresh in confidence_thresholds:
        for mode in modes:
            bankroll = starting_bankroll
            trades = []
            wins = 0
            losses = 0
            
            for i in range(window_size, len(candles) - window_size, window_size):
                window_ts = candles[i]["open_time"].timestamp()
                window_open_price = candles[i]["open"]
                
                current_price = candles[i + window_size - 1]["close"]
                window_candles = candles[max(0, i - 30):i]
                
                result = analyze(current_price, window_open_price, window_candles)
                
                confidence = result["confidence"]
                if confidence < conf_thresh:
                    confidence = conf_thresh
                
                if mode == "safe":
                    bet_size = bankroll * 0.25
                elif mode == "aggressive":
                    if bankroll >= starting_bankroll:
                        bet_size = bankroll - starting_bankroll
                    else:
                        bet_size = min_bet
                else:
                    bet_size = bankroll
                
                bet_size = max(bet_size, min_bet)
                
                delta_pct = ((current_price - window_open_price) / window_open_price) * 100
                token_price = simulate_token_price(delta_pct)
                
                shares = floor(bet_size / token_price)
                if shares < 5:
                    shares = 5
                    bet_size = shares * token_price
                
                direction = result["direction"]
                actual_direction = 1 if delta_pct > 0 else -1
                
                win = direction == actual_direction
                if win:
                    wins += 1
                    pnl = bet_size * (1.0 / token_price - 1)
                    bankroll += pnl
                else:
                    losses += 1
                    bankroll -= bet_size
                
                trades.append({
                    "window_ts": window_ts,
                    "direction": "up" if direction == 1 else "down",
                    "conf": confidence,
                    "bet": bet_size,
                    "token_price": token_price,
                    "win": win,
                    "pnl": pnl if win else -bet_size
                })
            
            total_trades = wins + losses
            win_rate = wins / total_trades if total_trades > 0 else 0
            total_pnl = bankroll - starting_bankroll
            
            results.append(BacktestResult(
                config=f"{mode}_{conf_thresh}",
                confidence_threshold=conf_thresh,
                mode=mode,
                total_trades=total_trades,
                wins=wins,
                losses=losses,
                win_rate=win_rate,
                final_bankroll=bankroll,
                total_pnl=total_pnl,
                max_drawdown=0,
                trades=trades
            ))
    
    return results


def create_excel_dashboard(results: List[BacktestResult], output_path: str):
    wb = openpyxl.Workbook()
    
    wb.remove(wb.active)
    
    ws_summary = wb.create_sheet("Summary")
    _create_summary_sheet(ws_summary, results)
    
    ws_trades = wb.create_sheet("Best Config Trades")
    _create_trades_sheet(ws_trades, results)
    
    ws_chart = wb.create_sheet("Bankroll Curves")
    _create_chart_sheet(ws_chart, results)
    
    wb.save(output_path)
    print(f"Dashboard saved to {output_path}")


def _create_summary_sheet(ws, results: List[BacktestResult]):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = ["Config", "Mode", "Conf Threshold", "Total Trades", "Wins", "Losses", "Win Rate", "Final Bankroll", "Total PnL", "ROI%"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    results_sorted = sorted(results, key=lambda x: x.final_bankroll, reverse=True)[:27]
    
    for row, result in enumerate(results_sorted, 2):
        ws.cell(row=row, column=1).value = result.config
        ws.cell(row=row, column=2).value = result.mode
        ws.cell(row=row, column=3).value = result.confidence_threshold
        ws.cell(row=row, column=4).value = result.total_trades
        ws.cell(row=row, column=5).value = result.wins
        ws.cell(row=row, column=6).value = result.losses
        ws.cell(row=row, column=7).value = f"{result.win_rate:.1%}"
        ws.cell(row=row, column=8).value = f"${result.final_bankroll:.2f}"
        ws.cell(row=row, column=9).value = f"${result.total_pnl:.2f}"
        ws.cell(row=row, column=10).value = f"{result.total_pnl/1.0*100:.1f}%"
        
        if result.final_bankroll > 1.5:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif result.final_bankroll < 0.5:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            fill = None
        
        for col in range(1, 11):
            if fill:
                ws.cell(row=row, column=col).fill = fill
    
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    ws.column_dimensions["A"].width = 18


def _create_trades_sheet(ws, results: List[BacktestResult]):
    if not results:
        return
    
    best = max(results, key=lambda x: x.final_bankroll)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    headers = ["#", "Window", "Direction", "Confidence", "Bet", "Token Price", "Win", "PnL", "Bankroll"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    bankroll = 1.0
    for i, trade in enumerate(best.trades[:100], 1):
        bankroll += trade["pnl"]
        
        ws.cell(row=i+1, column=1).value = i
        ws.cell(row=i+1, column=2).value = datetime.fromtimestamp(trade["window_ts"]).strftime("%m/%d %H:%M")
        ws.cell(row=i+1, column=3).value = trade["direction"]
        ws.cell(row=i+1, column=4).value = f"{trade['conf']:.0%}"
        ws.cell(row=i+1, column=5).value = f"${trade['bet']:.2f}"
        ws.cell(row=i+1, column=6).value = f"${trade['token_price']:.2f}"
        ws.cell(row=i+1, column=7).value = "YES" if trade["win"] else "NO"
        ws.cell(row=i+1, column=8).value = f"${trade['pnl']:.2f}"
        ws.cell(row=i+1, column=9).value = f"${bankroll:.2f}"
        
        if trade["win"]:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col in range(1, 10):
            ws.cell(row=i+1, column=col).fill = fill
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _create_chart_sheet(ws, results: List[BacktestResult]):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    ws.cell(row=1, column=1).value = "Trade #"
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    
    best_configs = sorted(results, key=lambda x: x.final_bankroll, reverse=True)[:5]
    
    for col, result in enumerate(best_configs, 2):
        ws.cell(row=1, column=col).value = result.config
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
    
    max_trades = max(len(r.trades) for r in best_configs)
    bankroll = 1.0
    
    for i in range(max_trades):
        ws.cell(row=i+2, column=1).value = i + 1
        
        for col, result in enumerate(best_configs, 2):
            if i < len(result.trades):
                bankroll += result.trades[i]["pnl"]
                ws.cell(row=i+2, column=col).value = round(bankroll, 2)
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14


def main():
    parser = argparse.ArgumentParser(description="Backtest comparison tool")
    parser.add_argument("--hours", type=int, default=72, help="Hours to backtest")
    parser.add_argument("--output", default="backtest_results.xlsx", help="Output file")
    args = parser.parse_args()
    
    candles = fetch_historical_data(args.hours)
    
    confidence_thresholds = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]
    modes = ["flat", "safe", "aggressive"]
    
    print(f"Running {len(confidence_thresholds) * len(modes)} configurations...")
    results = run_backtest(candles, confidence_thresholds, modes)
    
    create_excel_dashboard(results, args.output)


if __name__ == "__main__":
    main()